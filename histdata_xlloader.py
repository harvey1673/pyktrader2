# -*- coding: utf-8 -*-
"""
Created on Mon Aug 21 09:58:32 2017

This function reads in data from excel file PriceStore to Database file Database.db 
The extension of the source file should be xlsx

The function has two options: either append the most recent entries (implemented, test required)
or refresh the whole database 

Assume all future prices are close prices unless otherwise stated (STL)
"""

import time
import sys
import sqlite3
import openpyxl

PRICESTORE_DATAFILE = "S:\Metals Risk Team\Metals_risk_data\PriceStore\PriceStore.xlsb"
MYSTEEL_DATAFILE = ""
BILLET_DATAFILE = {"C:/Users/j291414/Desktop/Prices.xlsx":{"Dom-Exp Spread":['cn_billet', 1, 50], "DailyIndexes":['bs_billet', 3, 4]},
                   "C:/Users/j291414/Desktop/SEA EAF.xlsb.xlsx":{"Margin":['sea_billet', 1, 4]}
                   }

list_futures = {
        'SHFE.RBCNY.Store': ['rb', 'SHFE', 'close'],
        'SHFE.HRCCNY.Store':['hc', 'SHFE', 'close'],
        'DCE.IOECNY.Store':['i', 'DCE', 'close'],
        'DCE.CKCCNY.Store':['jm', 'DCE', 'close'],
        'SGX.IO.Store':['FE', 'SGX', 'close'],
        'SGX.HRC.Store':['hc','SGX', 'close'],
#        'SHFE.RBCNY.STL.Store': ['RB', 'CNY', 'SHFE', 'Settle'],
#        'SHFE.HRCCNY.STL.Store':['HRC', 'CNY', 'SHFE', 'Settle'],
#        'DCE.IOECNY.STL.Store':['IOE', 'CNY', 'DCE', 'Settle']
        }

# list of spot prices
list_spot = {
        'PLATTS.BLSPOT.Store':[['plt_bl', 2],],
        'TSI.SCRAPSPOT.Store':[['tsi_scrap', 2],],
        'PLATTS.IOSPOT.Store':[['plt_io62', 3],['plt_io65',4]],
        'PLATTS.CNSPOT.Store':[['plt_cn', 2],],
        'TSI.IOSPOT.Store':[['tis_io62', 3],]
        }

# list of port prices and Mandarin translation
dct = {u'FMG\u6df7\u5408\u7c89': ['FMG Fine', 'fmgF'],
       u'PB\u5757': ['PB Lump', 'pbL'],
         u'PB\u7c89': ['PB Fine', 'pbF'],
         u'\u5357\u975e\u5757': ['South Africa Lump', 'saL'],
         u'\u5357\u975e\u7c89': ['South Africa Fine', 'saF'],
         u'\u5361\u62c9\u52a0\u65af\u7c89': ['Carajas Fine', 'crjF'],
         u'\u5370\u5ea6\u5757\u77ff': ['India Lump', 'inL'] ,
         u'\u5370\u5ea6\u7c89\u77ff': ['India Fine', 'inF'],
         u'\u5df4\u5757': ['Brazil Mixed', 'brM'],
         u'\u5df4\u6df7': ['Brazil Mixed', 'brM'],
         u'\u5df4\u7c97': ['Brazil Thick', 'brT'],
         u'\u5df4\u897f\u5757': ['Brazil Lump', 'brL'],
         u'\u5df4\u897f\u5757\u77ff': ['Brazil Lump', 'brL'],
         u'\u666e\u5229\u7c89': ['Pro Lead Fine', 'plF'],
         u'\u6768\u8fea\u7c89': ['Yandi Fine', 'ydF'],
         u'\u7ebd\u66fc\u5757': ['Newman Lump', 'nmL'],
         u'\u7ebd\u66fc\u7c89': ['Newman Fine', 'nmF'],
         u'\u7f57\u5e03\u6cb3\u5757': ['Robe River Lump', 'rrL'],
         u'\u7f57\u5e03\u6cb3\u7c89': ['Robe River Fine', 'rrF'],
         u'\u8d85\u7279\u7c89': ['Super Special Fine', 'ssF'],
         u'\u91d1\u5e03\u5df4\u7c89': ['Jimblebar Fine', 'jbF'],
         u'\u963f\u7279\u62c9\u65af\u7c89': ['Atlas Fine', 'atlF'],
         u'\u9ad8\u7845\u5df4\u7c97': ['High Silica Brazil Thick', 'hsbrT'],
         u'\u9ea6\u514b\u7c89': ['MAC Fine', 'macF'],
         u'\u5df4\u7c97\uff08SSFG\uff09': ['Brazil Thick (SSFG)', 'ssfg'],
         u'\u5df4\u6df7\uff08BRBF\uff09': ['Brazil Mixed (BRBF)', 'brbf'],
         u'\u5df4\u7c97\uff08SFNG\uff09': ['Brazil Thick (SFNG)', 'sfng'],
         u'\u6df7\u5408\u7c89': ['Mixed Fine', 'mixF'],
         u'\u5df4\u7c97 (SFNG)': ['Brazil Thick (SSFG)', 'ssfg'],
         u'\u5df4\u7c97(SFNG)': ['Brazil Thick (SSFG)', 'ssfg']
         }

ports = ['RIZHAO', 'JINGTANG', 'JIANGYIN', 'LIANYUN', 'LANSHAN', 'QINGDAO', 
             'CAOFEIDIAN', 'TIANJIN', 'TAICANG']
list_port = {
        'MYSTEELRIZHAOIO.Store':['RIZHAO', 'rz'],
        'MYSTEELJINGTANGIO.Store':['JINGTANG', 'jt'],
        'MYSTEELJIANGYINIO.Store':['JIANGYIN', 'jy'],
        'MYSTEELLIANYUNIO.Store':['LIANYUN', 'ly'],
        'MYSTEELLANSHANIO.Store':['LANSHAN', 'ls'],
        'MYSTEELQINGDAOIO.Store':['QINGDAO', 'qd'],
        'MYSTEELCAOFEIDIANIO.Store':['CAOFEIDIAN', 'cfd'],
        'MYSTEELTIANJINIO.Store':['TIANJIN', 'tj'],
        'MYSTEELTAICANG.Store':['TAICANG', 'tc']}

'''
These are some handy functions
'''

def futures_entry_exist(db_cursor, instID, exch, date):
    cursor = db_cursor.execute("SELECT COUNT(*) FROM fut_daily WHERE instID = ? AND exch = ? AND date = ?",
                       (instID, exch, date))
    
    if cursor.fetchone()[0] == 0:
        return False
    else:
        return True
    
def update_future_entry(db_cursor, instID, exch, date, field, price):
    db_cursor.execute("UPDATE fut_daily \
              SET {} = ? \
              WHERE instID = ? AND exch = ? AND date = ?".format(field),
              (price, instID, exch, date))
    return

def insert_future_entry(db_cursor, instID, exch, date, field, price):
    db_cursor.execute("INSERT INTO fut_daily (instID, exch, date, {}) VALUES (?, ?, ?, ?)".format(field),
              (instID, exch, date, price))
    return       
    
def read_futures_maturity(ws):
    maturity = []
    for row in ws.iter_rows(max_row = 1, min_col = 2):
        for cell in row:
            maturity.append(cell.value)
    maturity = filter(None, maturity)
    return maturity

def read_steel_code(ws, num_row):
    ret = []
    for row in ws.iter_rows(min_row = 2, max_row=num_row+1, min_col = 8, max_col=8):
        for cell in row:
            ret.append(cell.value)
    return ret

def read_steel_date(ws):
    ret = []
    for row in ws.iter_rows(min_row = 2, min_col = 1, max_col=1):
        for cell in row:
            ret.append(cell.value)
    ret = filter(None, ret)
    return ret

def read_date(ws, refresh, refresh_lines):
    date = []
    if not refresh:
        for col in ws.iter_cols(min_row = 2, max_col = 1):
            for cell in col:
                date.append(cell.value)
        date = filter(None, date)
    else:
        for col in ws.iter_rows(min_row = 2, max_row = refresh_lines + 2, max_col = 1):
            for cell in col:
                date.append(cell.value)
        date = filter(None, date)
    return date

def read_futures_prices(ws, num_row, i):
    price = []
    for row in ws.iter_rows(min_row = 2, max_row = num_row+1, min_col = i+2, max_col = i+2):
        for cell in row:
            price.append(cell.value)
    return price

def read_spot_prices(ws, price_col, num_date):
    price = []
    for row in ws.iter_rows(min_row=2, max_row=num_date+1, min_col=price_col, max_col=price_col):
        for cell in row:
            price.append(cell.value)
    return price

def read_port_prices(ws, num_row, i):
    price = []
    for row in ws.iter_rows(min_row=3, max_row = num_row+2, min_col=i+2, max_col=i+2):
        for cell in row:
            price.append(cell.value)
    return price

def read_steel_price(ws, num_row):
    ret = [] 
    for row in ws.iter_rows(min_row = 2, max_row = num_row+1, min_col = 5, max_col=5):
        for cell in row:
            ret.append(cell.value)
    return ret

def translate(ch, fe):
    return dct[ch][1] + '_' + str(int(round(100*fe)))
    
def prod_to_name(ws):
    prod = []    
    for row in ws.iter_rows(max_row = 1, min_col = 2):
        for cell in row:
            prod.append(cell.value)
    prod = filter(None, prod)
    fe = []
    for col in ws.iter_cols(min_row = 2, max_row = 2, min_col = 2):
        for cell in col:
            fe.append(cell.value)
    fe = filter(None, fe)
    
    if len(prod) != len(fe):
        print('length of product name and fe contain not equal')
    
    name = []
    i = 0
    for i in range(len(prod)):
        name.append(translate(prod[i], fe[i]))
    return name

def get_spotID(ws):
    worksheet = ws.title
    port = list_port[worksheet][1]
    name = prod_to_name(ws)
    spotID = []
    i = 0
    for i in range(len(name)):
        spotID.append(port + '_' + name[i])
    return spotID


def get_instID(maturity, product):
    inst = []
    for d in maturity:
        year = d.strftime('%Y')[2:]
        month = d.strftime('%m')
        name = product + year + month
        inst.append(name)
    return inst

def delete_null(spotID, date, price):
    i = 0
    index = []
    for i in range(len(spotID)):
        if spotID[i] is None:
            index.append(i)
            
    for e in index:
        spotID.pop(e)
        date.pop(e)
        price.pop(e)
        
    return spotID, date, price

def read_billet_date(ws, date_col):
    date = []
    for col in ws.iter_cols(min_row = 2, min_col=date_col, max_col=date_col):
        for cell in col:
            date.append(cell.value)
    date = filter(None, date)
    return date
    
def read_billet_prices(ws, nEntry, col):
    price = [] 
    for col in ws.iter_rows(min_row=2, max_row = nEntry+1, min_col=col, max_col=col):
        for cell in col:
            price.append(cell.value)
    return price

def billet_to_db(cursor, spotID, date, prices):
    num = len(date)
    for i in range(num):
        if prices[i] is not None:
            cursor.execute("INSERT INTO spot_daily VALUES (?, ?, ?)", 
                           (spotID, date[i], prices[i]))


#+----------------------------------------------------------------------------+
''' loaders '''
#+----------------------------------------------------------------------------+
def pricestore_loader(db, refresh = False, refresh_lines = 0, \
						xlfile = PRICESTORE_DATAFILE,
						table_config = {'fut_daily': 'fut_daily', 'spot_daily': 'spot_daily'}):
    ''' input refresh as True if one wants to refresh recent entries and input 
        number of new lines as refresh_lines 
    '''
    start = time.time()
    
    print('\n')
    print('----------------------------------------------------------------')
    print('establishing connection')
    print('----------------------------------------------------------------')
    db_conn = sqlite3.connect(db)
    db_cursor = db_conn.cursor()
    wb = openpyxl.load_workbook(xlfile)
    
    
    ''' read in futures price'''
    print('\n')
    print('----------------------------------------------------------------')
    print('loading futures prices ')
    print('----------------------------------------------------------------')
    for work_sheet in list_futures:
        print('loading Sheet {}'.format(work_sheet))          
        ws = wb[work_sheet]
        maturity = read_futures_maturity(ws)
        date = read_date(ws, refresh, refresh_lines)
        prod = list_futures[work_sheet][0]
        exchange = list_futures[work_sheet][1]
        field = list_futures[work_sheet][2]
        instID = get_instID(maturity, prod)
        i = 0
        num_row = len(date)
        num_col = len(maturity)
        for i in range(num_col):
            price = read_futures_prices(ws, num_row, i)
            j = 0
            for j in range(num_row):
                #check existance
                if futures_entry_exist(db_cursor, instID[i], exchange, date[j]):
                    update_future_entry(db_cursor, instID[i], exchange, date[j], field, price[j])
                else:
                    insert_future_entry(db_cursor, instID[i], exchange, date[j], field, price[j])
        
        #clear up Null entries
        db_cursor.execute("DELETE FROM fut_daily \
                  WHERE open IS NULL AND close IS NULL AND high IS NULL AND low IS NULL \
                  AND volume IS NULL AND openInterest IS NULL")
        db_conn.commit()    
    print('----------------------------------------------------------------')
    print('futures prices loading complete')
    print('----------------------------------------------------------------')
    
    
    ''' read in spot prices '''
    print('\n')
    print('----------------------------------------------------------------')
    print('loading spot prices ')
    print('----------------------------------------------------------------')
    for work_sheet in list_spot:
        for element in list_spot[work_sheet]:
            print('loading Sheet {}'.format(work_sheet))
            ws = wb[work_sheet]
            spotID = element[0]
            price_col = element[1]
            date = read_date(ws, refresh, refresh_lines)
            price = read_spot_prices(ws, price_col, len(date))
            
            i = 0
            for i in range(len(date)):
                db_cursor.execute("INSERT INTO spot_daily VALUES (?, ?, ?)", (spotID, date[i], price[i]))
            db_conn.commit()
    print('----------------------------------------------------------------')
    print('spot prices loading complete')
    print('----------------------------------------------------------------')
    
    
    ''' read in port prices '''
    print('\n')
    print('----------------------------------------------------------------')
    print('loading port prices ')
    print('----------------------------------------------------------------')
    for work_sheet in list_port:
        print('loading Sheet {}'.format(work_sheet))
        ws = wb[work_sheet]
        date = read_date(ws, refresh, refresh_lines)
        spotID = get_spotID(ws)
        
        i = 0
        num_row = len(date)
        num_col = len(spotID)
        for i in range(num_col):
            price = read_port_prices(ws, num_row, i)
            j = 0
            for j in range(num_row):
                db_cursor.execute("INSERT INTO spot_daily (spotID, date, price) VALUES (?, ?, ?)", 
                          (spotID[i], date[j], price[j]))
        db_conn.commit()
    db_conn.close()
    end = time.time()
    elapsed = end - start
    print('----------------------------------------------------------------')
    print('operation complete. run time = %3f seconds' % (elapsed,))
    print('----------------------------------------------------------------')


#+----------------------------------------------------------------------------+
def mysteel_loader(db, xlfile = MYSTEEL_DATAFILE, table_config = {'fut_daily': 'fut_daily', 'spot_daily': 'spot_daily'}):
    ''' read in steel spot prices '''
    start = time.time()
    print('\n')
    print('----------------------------------------------------------------')
    print('loading steel prices ')
    print('----------------------------------------------------------------')
    db_conn = sqlite3.connect(db)
    db_cursor = db_conn.cursor()
    wb_s = openpyxl.load_workbook(xlfile)
    ws = wb_s['Historical']
    date = read_steel_date(ws)
    spotID = read_steel_code(ws, len(date))
    price = read_steel_price(ws, len(date))
    
    if len(date)==len(price) and len(price)==len(spotID):
        print('all fine 1')
    else:
        print('data length not equal 1')
    
    spotID, date, price = delete_null(spotID, date, price)  
    
    if len(date)==len(price) and len(price)==len(spotID):
        print('all fine 2')
    else:
        print('data length not equal 2')
    
    j=0
    num_row = len(date)
    for j in range(num_row):
        db_cursor.execute("INSERT INTO spot_daily VALUES (?, ?, ?)",
                  (spotID[j], date[j], price[j]))
    
    db_conn.commit()
    print('----------------------------------------------------------------')
    print('steel prices loading complete')
    print('----------------------------------------------------------------')
    
    ''' operation end '''
    db_conn.close()
    end = time.time()
    elapsed = end - start
    print('\n')
    print('----------------------------------------------------------------')
    print('operation complete. run time = %3f seconds' % (elapsed,))
    print('----------------------------------------------------------------')
#------------------------------------------------------------------------------

def billet_loader(db, xlfile = BILLET_DATAFILE):
    ''' read in billet prices '''
    start = time.time()
    print('\n')
    print('----------------------------------------------------------------')
    print('loading billet prices ')
    print('----------------------------------------------------------------')
    db_conn = sqlite3.connect(db)
    db_cursor = db_conn.cursor()
    n = 1
    
    for workbook in xlfile: 
        print('loading Workbook {}'.format(n))
        n += 1
        wb_s = openpyxl.load_workbook(workbook, data_only=True)
        
        for worksheet in xlfile[workbook]:
            ws = wb_s[worksheet]
            spotID = xlfile[workbook][worksheet][0]
            date_col = xlfile[workbook][worksheet][1]
            data_col = xlfile[workbook][worksheet][2]
            
            date = read_billet_date(ws, date_col)
            billet_prices = read_billet_prices(ws, len(date), data_col)
            billet_to_db(db_cursor, spotID, date, billet_prices)
            
            print('\t{} loading complete'.format(worksheet))
            db_conn.commit()
            
    print('----------------------------------------------------------------')
    print('billet prices loading complete')
    print('----------------------------------------------------------------')
    
    ''' operation end '''
    db_conn.close()
    end = time.time()
    elapsed = end - start
    print('\n')
    print('----------------------------------------------------------------')
    print('operation complete. run time = %3f seconds' % (elapsed,))
    print('----------------------------------------------------------------')
